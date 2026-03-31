import sys
import argparse
from openpyxl import load_workbook

MAX_ROW = 1000
MAX_COL = 1000

def parse_excel_with_hh(xlsx_path, geometry=[0,0,1,1], minblank=1, workaround=[]):
    if "usebr" in workaround:
        br = "<br>\n"
    else:
        br = "  \n"

    wb = load_workbook(xlsx_path, data_only=True)
    for sheetname in wb.sheetnames:
        print(f"# {sheetname}")
        ws = wb[sheetname]
        data_col, data_row, header_width, header_height = geometry
        blank = minblank
        for row in range(data_row, MAX_ROW):
            term = True
            for hdr_col in range(data_col - header_width, data_col):
                hdr = ws.cell(row=row, column=hdr_col).value
                if hdr:
                    term = False
                    indent = hdr_col - (data_col - header_width) + 2
                    print("#"*indent + f" {hdr}")
            if term:
                if blank == 0:
                    break
                blank -= 1
            else:
                blank = minblank
                    
            for col in range(data_col, MAX_COL):
                value = ws.cell(row=row, column=col).value

                term = True
                line = ""
                for hdr_row in range(data_row - header_height, data_row):
                    hdr = ws.cell(row=hdr_row, column=col).value
                    if hdr:
                        term = False
                        indent = hdr_row - (data_row - header_height)
                        if hdr_row < data_row -1:
                            line += " "*indent*2 + "-" + f" {hdr}\n"
                        elif value or "dense" in workaround:
                            line += " "*indent*2 + "-" + f" {hdr}: "
                if term:
                    continue

                if value or "dense" in workaround:
                    print(f"{line}", end="")
                    value = str(value)
                    if "escmd" in workaround:
                        value = value.replace("---", "--")
                    value = br.join(
                        _ for _ in value.splitlines()
                        if _.strip()
                        )
                    print(f"{value}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("filename", help="file name")
    parser.add_argument("--geometry", type=str, default="2x2+1+1", help="XxY[+W+H]: default data (X=1,Y=1), header (W=1,H=1)")
    parser.add_argument("--dense", action="store_true", help="display if empty")
    parser.add_argument("--escmd", action="store_true", help="escape markdown")
    parser.add_argument("--usebr", action="store_true", help="use <br> rather two spaces")
    parser.add_argument("--minblank", type=int, default=1, help="minimum blank lines: default 1")
    args = parser.parse_args()

    tmp = args.geometry.split("+")
    position = [int(x) for x in tmp[0].split("x")]
    layer = [int(tmp[1]), int(tmp[2])]
    geometry = position + layer
    workaround = []
    if args.dense:
        workaround.append("dense")
    if args.escmd:
        workaround.append("escmd")
    if args.usebr:
        workaround.append("usebr")
    if position[0] < layer[0] or position[1] < layer[1]:
        print(f"geometry mismatch: {geometry}", file=sys.stderr)
        sys.exit()

    parse_excel_with_hh(args.filename, geometry, args.minblank, workaround)
